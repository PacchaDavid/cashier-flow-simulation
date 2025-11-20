"""
simulacion_cajas.py
Simulación discreta de cajas de supermercado para la práctica:
'¿Cuándo abrir una nueva caja? (enfoque negocio)'

Genera:
- matriz_corridas.xlsx  (detalle réplicas)
- resumen_configuraciones.xlsx (resumen agregado por s)
- replicas_detalle.csv
- resumen_configuraciones.csv
- CT_vs_s.png, SLA_vs_s.png, rho_vs_s.png

Uso: python simulacion_cajas.py
Requiere: numpy, pandas, matplotlib, openpyxl
"""
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from dataclasses import dataclass
from typing import List, Optional
import os

# -----------------------
# PARÁMETROS
# -----------------------
@dataclass
class Parametros:
    # Tasas
    lambda_llegadas: float = 0.8      # clientes/minuto (48 clientes/h)
    mu_servicio: float = 0.35         # servicios/minuto (1/2.86)
    # Costos (USD)
    costo_caja: float = 0.50          # USD/minuto por caja
    costo_espera: float = 0.30        # USD/minuto por cliente (tiempo en sistema)
    costo_sla: float = 5.00           # USD por punto porcentual de incumplimiento
    # SLA
    sla_tiempo_max: float = 8.0       # minutos
    sla_porcentaje_min: float = 80.0  # %
    # Simulación
    tiempo_simulacion: int = 480      # minutos (8 horas)
    num_replicas: int = 30            # Réplicas por configuración (>=10 recomendado)
    seed_base: int = 1000             # semilla base para reproducibilidad
    # Producción de archivos
    output_dir: str = "resultados_sim"

# -----------------------
# CLASE CLIENTE
# -----------------------
class Cliente:
    contador = 0
    def __init__(self, tiempo_llegada: float):
        Cliente.contador += 1
        self.id = Cliente.contador
        self.tiempo_llegada = tiempo_llegada
        self.tiempo_inicio_servicio = None
        self.tiempo_salida = None
    def tiempo_en_sistema(self) -> Optional[float]:
        if self.tiempo_salida is None:
            return None
        return self.tiempo_salida - self.tiempo_llegada
    def tiempo_espera(self) -> Optional[float]:
        if self.tiempo_inicio_servicio is None:
            return None
        return self.tiempo_inicio_servicio - self.tiempo_llegada

# -----------------------
# SIMULADOR
# -----------------------
class SimuladorSupermercado:
    def __init__(self, num_cajas: int, params: Parametros):
        self.num_cajas = num_cajas
        self.params = params
        self._init_state()
    def _init_state(self):
        self.tiempo_actual = 0.0
        self.cola: List[Cliente] = []
        self.cajas_ocupadas: List[Optional[Cliente]] = [None] * self.num_cajas
        self.tiempos_fin_servicio: List[float] = [float('inf')] * self.num_cajas
        self.clientes_atendidos: List[Cliente] = []
        # tiempo hasta la próxima llegada (intervalo)
        self.tiempo_proximo_intervalo = np.random.exponential(1 / self.params.lambda_llegadas)
    def _generar_tiempo_llegada(self) -> float:
        return np.random.exponential(1 / self.params.lambda_llegadas)
    def _generar_tiempo_servicio(self) -> float:
        return np.random.exponential(1 / self.params.mu_servicio)
    def _asignar_cliente_a_caja(self, cliente: Cliente) -> bool:
        for i in range(self.num_cajas):
            if self.cajas_ocupadas[i] is None:
                self.cajas_ocupadas[i] = cliente
                cliente.tiempo_inicio_servicio = self.tiempo_actual
                tiempo_servicio = self._generar_tiempo_servicio()
                self.tiempos_fin_servicio[i] = self.tiempo_actual + tiempo_servicio
                return True
        return False
    def simular_una_replica(self):
        # Reiniciar estado para la réplica
        self._init_state()
        while self.tiempo_actual < self.params.tiempo_simulacion:
            tiempo_proxima_llegada = self.tiempo_actual + self.tiempo_proximo_intervalo
            tiempo_proxima_salida = min(self.tiempos_fin_servicio)
            # Evento más cercano
            if tiempo_proxima_llegada < tiempo_proxima_salida:
                # LLEGADA
                self.tiempo_actual = tiempo_proxima_llegada
                if self.tiempo_actual >= self.params.tiempo_simulacion:
                    break
                cliente = Cliente(self.tiempo_actual)
                if not self._asignar_cliente_a_caja(cliente):
                    self.cola.append(cliente)
                self.tiempo_proximo_intervalo = self._generar_tiempo_llegada()
            else:
                # SALIDA
                self.tiempo_actual = tiempo_proxima_salida
                if self.tiempo_actual >= self.params.tiempo_simulacion:
                    break
                caja_idx = self.tiempos_fin_servicio.index(tiempo_proxima_salida)
                cliente_saliendo = self.cajas_ocupadas[caja_idx]
                if cliente_saliendo is None:
                    # Defensa: limpiar y continuar
                    self.tiempos_fin_servicio[caja_idx] = float('inf')
                    self.cajas_ocupadas[caja_idx] = None
                    continue
                cliente_saliendo.tiempo_salida = self.tiempo_actual
                self.clientes_atendidos.append(cliente_saliendo)
                # Liberar caja
                self.cajas_ocupadas[caja_idx] = None
                self.tiempos_fin_servicio[caja_idx] = float('inf')
                # Si hay cola, atender siguiente
                if len(self.cola) > 0:
                    siguiente = self.cola.pop(0)
                    self._asignar_cliente_a_caja(siguiente)
        return self._calcular_metricas()
    def _calcular_metricas(self):
        if len(self.clientes_atendidos) == 0:
            return None
        tiempos_sistema = [c.tiempo_en_sistema() for c in self.clientes_atendidos]
        tiempos_espera = [c.tiempo_espera() for c in self.clientes_atendidos if c.tiempo_espera() is not None]
        tiempo_promedio_sistema = float(np.mean(tiempos_sistema))
        tiempo_promedio_espera = float(np.mean(tiempos_espera)) if len(tiempos_espera) > 0 else 0.0
        cumple_sla = sum(1 for t in tiempos_sistema if t <= self.params.sla_tiempo_max)
        porcentaje_sla = (cumple_sla / len(tiempos_sistema)) * 100.0
        costo_cajas = self.num_cajas * self.params.costo_caja * self.params.tiempo_simulacion
        costo_espera_total = float(sum(tiempos_sistema) * self.params.costo_espera)
        incumplimiento_sla = max(0.0, self.params.sla_porcentaje_min - porcentaje_sla)
        penalizacion_sla = incumplimiento_sla * self.params.costo_sla
        costo_total = costo_cajas + costo_espera_total + penalizacion_sla
        utilizacion = self.params.lambda_llegadas / (self.num_cajas * self.params.mu_servicio)
        return {
            'num_cajas': self.num_cajas,
            'clientes_atendidos': len(self.clientes_atendidos),
            'tiempo_promedio_sistema': tiempo_promedio_sistema,
            'tiempo_promedio_espera': tiempo_promedio_espera,
            'porcentaje_sla': porcentaje_sla,
            'utilizacion': utilizacion,
            'costo_cajas': costo_cajas,
            'costo_espera': costo_espera_total,
            'penalizacion_sla': penalizacion_sla,
            'costo_total': costo_total
        }

# -----------------------
# FUNCIONES DE PRESENTACIÓN
# -----------------------
def imprimir_encabezado():
    """Imprime encabezado profesional del experimento"""
    print("\n" + "="*80)
    print(" " * 15 + "PRÁCTICA 005: ¿CUÁNDO ABRIR UNA NUEVA CAJA?")
    print(" " * 20 + "Simulación de Eventos Discretos")
    print("="*80)

def imprimir_parametros(params: Parametros):
    """Imprime los parámetros de la simulación"""
    print("\n📊 PARÁMETROS DE LA SIMULACIÓN:")
    print("-" * 80)
    print(f"  • Tasa de llegadas (λ):        {params.lambda_llegadas} clientes/min ({params.lambda_llegadas*60:.0f} clientes/hora)")
    print(f"  • Tasa de servicio (μ):        {params.mu_servicio} servicios/min ({1/params.mu_servicio:.2f} min/cliente)")
    print(f"  • Tiempo de simulación:        {params.tiempo_simulacion} minutos ({params.tiempo_simulacion/60:.1f} horas)")
    print(f"  • Réplicas por configuración:  {params.num_replicas}")
    print(f"  • Costo por caja:              ${params.costo_caja:.2f}/min")
    print(f"  • Costo por espera:            ${params.costo_espera:.2f}/min/cliente")
    print(f"  • Penalización SLA:            ${params.costo_sla:.2f} por punto %")
    print(f"  • SLA objetivo:                ≥{params.sla_porcentaje_min:.0f}% con T≤{params.sla_tiempo_max:.0f} min")
    print("-" * 80)

def imprimir_progreso(s: int, total_configs: int, idx: int):
    """Imprime barra de progreso"""
    barra_len = 40
    progreso = (idx + 1) / total_configs
    bloques = int(barra_len * progreso)
    barra = "█" * bloques + "░" * (barra_len - bloques)
    print(f"\n🔄 Simulando s={s} cajas: [{barra}] {progreso*100:.0f}%")

def imprimir_matriz_muestra(df_all: pd.DataFrame, n_filas: int = 10):
    """Imprime una muestra de la matriz de corridas"""
    print("\n" + "="*80)
    print("📋 MATRIZ DE CORRIDAS (muestra de primeras {} filas)".format(n_filas))
    print("="*80)
    
    # Seleccionar columnas clave para mostrar
    cols_mostrar = ['num_cajas', 'replica', 'clientes_atendidos', 
                    'tiempo_promedio_sistema', 'porcentaje_sla', 
                    'utilizacion', 'costo_total']
    
    df_muestra = df_all[cols_mostrar].head(n_filas).copy()
    
    # Formatear columnas
    df_muestra['tiempo_promedio_sistema'] = df_muestra['tiempo_promedio_sistema'].round(2)
    df_muestra['porcentaje_sla'] = df_muestra['porcentaje_sla'].round(1)
    df_muestra['utilizacion'] = df_muestra['utilizacion'].round(3)
    df_muestra['costo_total'] = df_muestra['costo_total'].round(2)
    
    # Renombrar columnas para presentación
    df_muestra = df_muestra.rename(columns={
        'num_cajas': 'Cajas',
        'replica': 'Rep',
        'clientes_atendidos': 'Clientes',
        'tiempo_promedio_sistema': 'T_sist',
        'porcentaje_sla': '%SLA',
        'utilizacion': 'ρ',
        'costo_total': 'CT'
    })
    
    print(df_muestra.to_string(index=False))
    print(f"\n[Mostrando {n_filas} de {len(df_all)} réplicas totales]")
    print("-" * 80)

def imprimir_resumen_detallado(df_resumen: pd.DataFrame):
    """Imprime resumen estadístico detallado"""
    print("\n" + "="*80)
    print("📊 RESUMEN ESTADÍSTICO POR CONFIGURACIÓN")
    print("="*80)
    
    for _, row in df_resumen.iterrows():
        s = int(row['num_cajas'])
        cumple_sla = row['porcentaje_sla_mean'] >= 80.0
        simbolo_sla = "✅" if cumple_sla else "❌"
        
        print(f"\n┌─ CONFIGURACIÓN: {s} CAJAS {simbolo_sla}")
        print(f"│")
        print(f"│  Réplicas válidas:     {int(row['n_replicas_validas'])}")
        print(f"│  Clientes atendidos:   {row['clientes_atendidos_mean']:.1f} (promedio)")
        print(f"│")
        print(f"│  📏 Tiempos:")
        print(f"│     • T sistema:        {row['tiempo_promedio_sistema_mean']:.2f} ± {row['tiempo_promedio_sistema_std']:.2f} min")
        print(f"│")
        print(f"│  🎯 Nivel de Servicio:")
        print(f"│     • %SLA:             {row['porcentaje_sla_mean']:.1f}% ± {row['porcentaje_sla_std']:.1f}%")
        print(f"│     • Cumple objetivo:  {'SÍ (≥80%)' if cumple_sla else 'NO (<80%)'}")
        print(f"│")
        print(f"│  ⚙️  Utilización:")
        print(f"│     • ρ:                {row['utilizacion_mean']:.3f}")
        print(f"│     • Estado:           ", end="")
        if row['utilizacion_mean'] >= 1.0:
            print("INESTABLE (ρ≥1)")
        elif row['utilizacion_mean'] >= 0.9:
            print("SATURADO (ρ≥0.9)")
        elif row['utilizacion_mean'] >= 0.7:
            print("SALUDABLE (0.7≤ρ<0.9)")
        elif row['utilizacion_mean'] >= 0.5:
            print("HOLGADO (0.5≤ρ<0.7)")
        else:
            print("SOBRECAPACIDAD (ρ<0.5)")
        print(f"│")
        print(f"│  💰 Costos (USD):")
        print(f"│     • Costo cajas:      ${row['costo_cajas_mean']:.2f}")
        print(f"│     • Costo espera:     ${row['costo_espera_mean']:.2f}")
        print(f"│     • Penalización SLA: ${row['penalizacion_sla_mean']:.2f}")
        print(f"│     • COSTO TOTAL:      ${row['costo_total_mean']:.2f} ± ${row['costo_total_std']:.2f}")
        print(f"└─")
    
    print("\n" + "-"*80)

def imprimir_tabla_comparativa(df_resumen: pd.DataFrame):
    """Imprime tabla comparativa resumida"""
    print("\n" + "="*80)
    print("📈 TABLA COMPARATIVA FINAL")
    print("="*80)
    
    display_df = df_resumen[['num_cajas','n_replicas_validas',
                             'tiempo_promedio_sistema_mean','porcentaje_sla_mean',
                             'utilizacion_mean','costo_total_mean']].copy()
    
    display_df = display_df.rename(columns={
        'num_cajas':'Cajas',
        'n_replicas_validas':'Rép',
        'tiempo_promedio_sistema_mean':'T_sist (min)',
        'porcentaje_sla_mean':'%SLA',
        'utilizacion_mean':'ρ',
        'costo_total_mean':'CT (USD)'
    })
    
    # Formatear
    display_df['T_sist (min)'] = display_df['T_sist (min)'].round(2)
    display_df['%SLA'] = display_df['%SLA'].round(1)
    display_df['ρ'] = display_df['ρ'].round(3)
    display_df['CT (USD)'] = display_df['CT (USD)'].round(2)
    
    print(display_df.to_string(index=False))
    print("-" * 80)

def imprimir_analisis_decision(df_resumen: pd.DataFrame):
    """Imprime análisis y recomendación final"""
    print("\n" + "="*80)
    print("🎯 ANÁLISIS Y DECISIÓN")
    print("="*80)
    
    # Filtrar configuraciones que cumplen SLA
    viables = df_resumen[df_resumen['porcentaje_sla_mean'] >= 80.0]
    
    if viables.empty:
        print("\n⚠️  ADVERTENCIA: Ninguna configuración cumple el SLA mínimo del 80%")
        print("   Se requieren más cajas o mejorar la velocidad de servicio.")
        return
    
    # Encontrar el óptimo
    optimo = viables.loc[viables['costo_total_mean'].idxmin()]
    s_optimo = int(optimo['num_cajas'])
    ct_optimo = optimo['costo_total_mean']
    sla_optimo = optimo['porcentaje_sla_mean']
    rho_optimo = optimo['utilizacion_mean']
    
    print(f"\n✅ CONFIGURACIÓN ÓPTIMA: {s_optimo} CAJAS")
    print(f"   • Costo total:       ${ct_optimo:.2f}/turno")
    print(f"   • Cumplimiento SLA:  {sla_optimo:.1f}% (objetivo: ≥80%)")
    print(f"   • Utilización:       ρ = {rho_optimo:.3f}")
    
    # Comparación con otras opciones
    print(f"\n📊 Comparación con alternativas:")
    
    for _, row in df_resumen.iterrows():
        s = int(row['num_cajas'])
        ct = row['costo_total_mean']
        sla = row['porcentaje_sla_mean']
        
        if s == s_optimo:
            continue
        
        diff_ct = ct - ct_optimo
        diff_sla = sla - sla_optimo
        cumple = "✅" if sla >= 80 else "❌"
        
        print(f"\n   {s} cajas {cumple}:")
        print(f"      • CT: ${ct:.2f} ({'+' if diff_ct > 0 else ''}{diff_ct:.2f} vs óptimo, {diff_ct/ct_optimo*100:+.1f}%)")
        print(f"      • SLA: {sla:.1f}% ({'+' if diff_sla > 0 else ''}{diff_sla:.1f} puntos)")
        
        if s < s_optimo and sla < 80:
            print(f"      → RECHAZADO: No cumple SLA + sistema inestable")
        elif s < s_optimo:
            print(f"      → RECHAZADO: Ahorra ${-diff_ct:.2f} pero pierde en servicio")
        elif s > s_optimo:
            mejora_sla = diff_sla
            costo_extra = diff_ct
            if mejora_sla > 0:
                costo_por_punto = costo_extra / mejora_sla
                print(f"      → NO JUSTIFICADO: Cuesta ${costo_extra:.2f} más por solo {mejora_sla:.1f}% mejor SLA")
                print(f"         (${costo_por_punto:.2f} por punto porcentual)")
            else:
                print(f"      → SOBRECAPACIDAD: Costo mayor sin beneficio en SLA")
    
    print("\n" + "-"*80)
    
    # Recomendación
    print("\n💡 RECOMENDACIÓN:")
    print(f"   Operar con {s_optimo} cajas durante condiciones normales.")
    print(f"   Implementar regla dinámica para picos de demanda (ρ > 0.85 o Lq > 5).")
    print("\n" + "="*80)

# -----------------------
# FUNCIONES DE FORMATEO EXCEL
# -----------------------
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def guardar_excel_formateado_matriz(df: pd.DataFrame, ruta: str):
    """Guarda la matriz de corridas con formato profesional"""
    # Renombrar columnas
    df_formato = df.rename(columns={
        'num_cajas': 'Cajas',
        'replica': 'Réplica',
        'seed': 'Semilla',
        'config_cajas': 'Config',
        'clientes_atendidos': 'Clientes',
        'tiempo_promedio_sistema': 'T_sistema',
        'tiempo_promedio_espera': 'T_espera',
        'porcentaje_sla': '%SLA',
        'utilizacion': 'Utilización',
        'costo_cajas': 'C_cajas',
        'costo_espera': 'C_espera',
        'penalizacion_sla': 'P_SLA',
        'costo_total': 'Costo_Total'
    })
    
    # Redondear
    df_formato['T_sistema'] = df_formato['T_sistema'].round(2)
    df_formato['T_espera'] = df_formato['T_espera'].round(2)
    df_formato['%SLA'] = df_formato['%SLA'].round(1)
    df_formato['Utilización'] = df_formato['Utilización'].round(3)
    df_formato['C_cajas'] = df_formato['C_cajas'].round(2)
    df_formato['C_espera'] = df_formato['C_espera'].round(2)
    df_formato['P_SLA'] = df_formato['P_SLA'].round(2)
    df_formato['Costo_Total'] = df_formato['Costo_Total'].round(2)
    
    # Guardar
    with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
        df_formato.to_excel(writer, sheet_name='Matriz', index=False)
        
        # Formatear
        ws = writer.sheets['Matriz']
        
        # Estilos
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatear encabezados
        for col in range(1, len(df_formato.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Ajustar anchos de columna
        anchos = {
            'Cajas': 8,
            'Réplica': 10,
            'Semilla': 10,
            'Clientes': 10,
            'T_sistema': 12,
            'T_espera': 12,
            '%SLA': 10,
            'Utilización': 12,
            'C_cajas': 12,
            'C_espera': 12,
            'P_SLA': 10,
            'Costo_Total': 13
        }
        
        for idx, col in enumerate(df_formato.columns, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = anchos.get(col, 12)
        
        # Centrar datos
        for row in range(2, len(df_formato) + 2):
            for col in range(1, len(df_formato.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'

def guardar_excel_formateado_resumen(df: pd.DataFrame, ruta: str):
    """Guarda el resumen con formato profesional"""
    # Renombrar columnas
    df_formato = df.rename(columns={
        'num_cajas': 'Cajas',
        'n_replicas_validas': 'Réplicas',
        'clientes_atendidos_mean': 'Clientes_μ',
        'tiempo_promedio_sistema_mean': 'T_sist_μ',
        'tiempo_promedio_sistema_std': 'T_sist_σ',
        'porcentaje_sla_mean': '%SLA_μ',
        'porcentaje_sla_std': '%SLA_σ',
        'utilizacion_mean': 'ρ',
        'costo_total_mean': 'CT_μ',
        'costo_total_std': 'CT_σ',
        'costo_cajas_mean': 'C_cajas',
        'costo_espera_mean': 'C_espera',
        'penalizacion_sla_mean': 'P_SLA'
    })
    
    # Redondear
    df_formato['Clientes_μ'] = df_formato['Clientes_μ'].round(1)
    df_formato['T_sist_μ'] = df_formato['T_sist_μ'].round(2)
    df_formato['T_sist_σ'] = df_formato['T_sist_σ'].round(2)
    df_formato['%SLA_μ'] = df_formato['%SLA_μ'].round(1)
    df_formato['%SLA_σ'] = df_formato['%SLA_σ'].round(1)
    df_formato['ρ'] = df_formato['ρ'].round(3)
    df_formato['CT_μ'] = df_formato['CT_μ'].round(2)
    df_formato['CT_σ'] = df_formato['CT_σ'].round(2)
    df_formato['C_cajas'] = df_formato['C_cajas'].round(2)
    df_formato['C_espera'] = df_formato['C_espera'].round(2)
    df_formato['P_SLA'] = df_formato['P_SLA'].round(2)
    
    # Guardar
    with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
        df_formato.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Formatear
        ws = writer.sheets['Resumen']
        
        # Estilos
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        optimo_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        inviable_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatear encabezados
        for col in range(1, len(df_formato.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Ajustar anchos
        anchos = {
            'Cajas': 8,
            'Réplicas': 10,
            'Clientes_μ': 12,
            'T_sist_μ': 11,
            'T_sist_σ': 11,
            '%SLA_μ': 10,
            '%SLA_σ': 10,
            'ρ': 8,
            'CT_μ': 11,
            'CT_σ': 10,
            'C_cajas': 11,
            'C_espera': 11,
            'P_SLA': 10
        }
        
        for idx, col in enumerate(df_formato.columns, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = anchos.get(col, 12)
        
        # Encontrar fila óptima (mínimo CT entre los que cumplen SLA)
        viables = df_formato[df_formato['%SLA_μ'] >= 80.0]
        if not viables.empty:
            idx_optimo = viables['CT_μ'].idxmin()
            fila_optimo = idx_optimo + 2  # +2 porque Excel empieza en 1 y hay header
        else:
            fila_optimo = None
        
        # Formatear filas de datos
        for idx, row in df_formato.iterrows():
            fila_excel = idx + 2
            cumple_sla = row['%SLA_μ'] >= 80.0
            
            for col in range(1, len(df_formato.columns) + 1):
                cell = ws.cell(row=fila_excel, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Resaltar óptimo o inviable
                if fila_optimo and fila_excel == fila_optimo:
                    cell.fill = optimo_fill
                    cell.font = Font(bold=True)
                elif not cumple_sla:
                    cell.fill = inviable_fill
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'

# -----------------------
# EXPERIMENTO COMPLETO
# -----------------------
def ejecutar_experimento(configuraciones_cajas: List[int], params: Parametros):
    all_replicas = []
    resumen_list = []
    os.makedirs(params.output_dir, exist_ok=True)
    
    # Imprimir encabezado
    imprimir_encabezado()
    imprimir_parametros(params)
    
    print("\n🚀 INICIANDO EXPERIMENTO")
    print(f"   Configuraciones a evaluar: {configuraciones_cajas}")
    print(f"   Total de simulaciones: {len(configuraciones_cajas)} × {params.num_replicas} = {len(configuraciones_cajas) * params.num_replicas}")
    
    # Ejecutar simulaciones
    for idx, s in enumerate(configuraciones_cajas):
        imprimir_progreso(s, len(configuraciones_cajas), idx)
        
        replicas_ok = []
        for r in range(params.num_replicas):
            seed = params.seed_base + r
            np.random.seed(seed)
            simulador = SimuladorSupermercado(s, params)
            metricas = simulador.simular_una_replica()
            if metricas is None:
                continue
            metricas['replica'] = r
            metricas['seed'] = seed
            replicas_ok.append(metricas)
        
        df_reps = pd.DataFrame(replicas_ok)
        if df_reps.empty:
            print(f"   ⚠️  Ninguna réplica válida para s={s}. Se omite.")
            continue
        
        df_reps['config_cajas'] = s
        all_replicas.append(df_reps)
        
        # Resumen estadístico
        resumen = {
            'num_cajas': s,
            'n_replicas_validas': len(df_reps),
            'clientes_atendidos_mean': df_reps['clientes_atendidos'].mean(),
            'tiempo_promedio_sistema_mean': df_reps['tiempo_promedio_sistema'].mean(),
            'tiempo_promedio_sistema_std': df_reps['tiempo_promedio_sistema'].std(ddof=1),
            'porcentaje_sla_mean': df_reps['porcentaje_sla'].mean(),
            'porcentaje_sla_std': df_reps['porcentaje_sla'].std(ddof=1),
            'utilizacion_mean': df_reps['utilizacion'].mean(),
            'costo_total_mean': df_reps['costo_total'].mean(),
            'costo_total_std': df_reps['costo_total'].std(ddof=1),
            'costo_cajas_mean': df_reps['costo_cajas'].mean(),
            'costo_espera_mean': df_reps['costo_espera'].mean(),
            'penalizacion_sla_mean': df_reps['penalizacion_sla'].mean()
        }
        resumen_list.append(resumen)
        
        print(f"   ✓ Completado: {len(df_reps)} réplicas válidas")
    
    # Concatenar y guardar archivos
    if len(all_replicas) > 0:
        df_all = pd.concat(all_replicas, ignore_index=True)
        
        # Guardar Excel FORMATEADO
        excel_path = os.path.join(params.output_dir, "matriz_corridas.xlsx")
        guardar_excel_formateado_matriz(df_all, excel_path)
        
        df_all.to_csv(os.path.join(params.output_dir, "replicas_detalle.csv"), index=False)
        
        # Mostrar muestra de la matriz
        imprimir_matriz_muestra(df_all, n_filas=12)
    
    df_resumen = pd.DataFrame(resumen_list).sort_values('num_cajas')
    
    # Guardar resumen FORMATEADO
    resumen_path = os.path.join(params.output_dir, "resumen_configuraciones.xlsx")
    guardar_excel_formateado_resumen(df_resumen, resumen_path)
    
    df_resumen.to_csv(os.path.join(params.output_dir, "resumen_configuraciones.csv"), index=False)
    
    # Mostrar resultados
    if not df_resumen.empty:
        imprimir_resumen_detallado(df_resumen)
        imprimir_tabla_comparativa(df_resumen)
        imprimir_analisis_decision(df_resumen)
        graficar_resumen(df_resumen, params.output_dir)
    else:
        print("\n⚠️  No hay configuraciones válidas. Revisa parámetros y réplicas.")
    
    # Resumen de archivos generados
    print("\n📁 ARCHIVOS GENERADOS:")
    print(f"   Directorio: {os.path.abspath(params.output_dir)}")
    print(f"   • matriz_corridas.xlsx")
    print(f"   • resumen_configuraciones.xlsx")
    print(f"   • replicas_detalle.csv")
    print(f"   • resumen_configuraciones.csv")
    print(f"   • CT_vs_s.png")
    print(f"   • SLA_vs_s.png")
    print(f"   • rho_vs_s.png")
    print("\n" + "="*80)
    print("✅ EXPERIMENTO COMPLETADO CON ÉXITO")
    print("="*80 + "\n")
    
    return df_resumen

def graficar_resumen(df_resumen: pd.DataFrame, output_dir: str):
    df = df_resumen.sort_values('num_cajas')
    
    print("\n🎨 Generando gráficos...")
    
    # CT vs s
    plt.figure(figsize=(7,4.5))
    plt.errorbar(df['num_cajas'], df['costo_total_mean'], yerr=df['costo_total_std'], 
                 marker='o', capsize=5, linewidth=2, markersize=8)
    plt.xlabel('Número de cajas', fontsize=12)
    plt.ylabel('Costo total medio (USD)', fontsize=12)
    plt.title('Costo total vs número de cajas', fontsize=14, fontweight='bold')
    plt.grid(alpha=0.3)
    plt.xticks(df['num_cajas'])
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "CT_vs_s.png"), dpi=300)
    plt.close()
    print("   ✓ CT_vs_s.png")
    
    # SLA vs s
    plt.figure(figsize=(7,4.5))
    plt.errorbar(df['num_cajas'], df['porcentaje_sla_mean'], yerr=df['porcentaje_sla_std'], 
                 marker='s', capsize=5, linewidth=2, markersize=8)
    plt.axhline(y=80, color='red', linestyle='--', linewidth=2, label='SLA mínimo (80%)')
    plt.xlabel('Número de cajas', fontsize=12)
    plt.ylabel('% clientes con T ≤ 8 min', fontsize=12)
    plt.title('Cumplimiento SLA vs número de cajas', fontsize=14, fontweight='bold')
    plt.legend(fontsize=10)
    plt.grid(alpha=0.3)
    plt.xticks(df['num_cajas'])
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "SLA_vs_s.png"), dpi=300)
    plt.close()
    print("   ✓ SLA_vs_s.png")
    
    # rho vs s
    plt.figure(figsize=(7,4.5))
    plt.plot(df['num_cajas'], df['utilizacion_mean'], marker='^', 
             linewidth=2, markersize=10)
    plt.axhline(y=1.0, color='red', linestyle='--', linewidth=2, label='Límite estabilidad (ρ=1)')
    plt.xlabel('Número de cajas', fontsize=12)
    plt.ylabel('Utilización media ρ', fontsize=12)
    plt.title('Utilización (ρ) vs número de cajas', fontsize=14, fontweight='bold')
    plt.legend(fontsize=10)
    plt.grid(alpha=0.3)
    plt.xticks(df['num_cajas'])
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "rho_vs_s.png"), dpi=300)
    plt.close()
    print("   ✓ rho_vs_s.png")


def analisis_sensibilidad(base_configuraciones: List[int], params: Parametros, variaciones: List[float]):
    """
    Ejecuta experimentos variando lambda y mu multiplicativamente por cada factor en variaciones.
    Guarda resultados por cada escenario en subcarpetas dentro de params.output_dir.
    """
    print("\n" + "="*80)
    print("🔬 ANÁLISIS DE SENSIBILIDAD")
    print("="*80)
    
    resultados = {}
    for factor in variaciones:
        print(f"\nVariación: {factor*100:.0f}% del valor base")
        
        # Variar lambda
        params_var = Parametros(**vars(params))
        params_var.lambda_llegadas = params.lambda_llegadas * factor
        params_var.output_dir = os.path.join(params.output_dir, f"sens_lambda_{int(factor*100)}")
        print(f"  • λ = {params_var.lambda_llegadas:.2f} clientes/min")
        ejecutar_experimento(base_configuraciones, params_var)
        
        # Variar mu
        params_var2 = Parametros(**vars(params))
        params_var2.mu_servicio = params.mu_servicio * factor
        params_var2.output_dir = os.path.join(params.output_dir, f"sens_mu_{int(factor*100)}")
        print(f"  • μ = {params_var2.mu_servicio:.2f} servicios/min")
        ejecutar_experimento(base_configuraciones, params_var2)
    print("="*80)

# -----------------------
# EJECUCIÓN PRINCIPAL
# -----------------------
if __name__ == "__main__":
    params = Parametros()
    configuraciones = [2, 3, 4, 5]
    
    # Ejecutar experimento principal
    df_res = ejecutar_experimento(configuraciones, params)
    
